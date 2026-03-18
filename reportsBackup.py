"""
Reports Module - UDLMS (UNIFIED)
ALL reports in one place: FD, RD, Daily, Monthly Collection, Weekly Collection
"""
from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for, session
from datetime import datetime
import logging
import os
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from db import get_db_connection
from login import login_required

logger = logging.getLogger(__name__)

# Try importing PDF generators
try:
    from your_pdf_generators import (
        generate_fdcollections,
        generate_rd_report,
        generate_daily_report,
        generate_fd_report,
        generate_rd_monthly_report,
        generate_fd_interest_report,
        generate_glance_report,
        generate_memberwise_report,
        generate_monthly_collection_pdf,
        generate_weekly_collection_pdf,
    )
    PDF_AVAILABLE = True
except ImportError as e:
    logger.warning(f"PDF generators not available: {e}")
    PDF_AVAILABLE = False

reports_bp = Blueprint("reports", __name__, url_prefix="/reports")


# -------------------------------------------------------------
# SQL Helper
# -------------------------------------------------------------
def fetch_data(query, params):
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        return columns, rows
    except Exception as e:
        logger.error(f"fetch_data error: {e}")
        raise
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


# -------------------------------------------------------------
# Excel Generator for Collection Reports
# -------------------------------------------------------------
def generate_excel(data, due_date, branch_name, center_name, filepath, is_weekly=False):
    """Generate Excel for collection reports - Blue (monthly) or Green (weekly)."""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Collection" if is_weekly else "Monthly Collection"
    
    # Title color
    title_color = "059669" if is_weekly else "0c2a78"
    title_text = "WEEKLY COLLECTION SHEET" if is_weekly else "MONTHLY COLLECTION SHEET"
    
    ws.merge_cells('A1:O1')
    ws['A1'] = f"{branch_name} - {title_text}"
    ws['A1'].font = Font(size=16, bold=True, color=title_color)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    ws.merge_cells('A2:O2')
    ws['A2'] = f"Due Date: {datetime.strptime(due_date, '%Y-%m-%d').strftime('%d-%b-%Y')} | Center: {center_name}"
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = ['S.No', 'Center', 'Member Code', 'Member Name', 'Mobile', 'Loan ID',
               'Principal Due', 'Interest Due', 'Savings Due', 'Total Due',
               'Principal Paid', 'Interest Paid', 'Savings Paid', 'Total Paid', 'Status']
    
    # Header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color=title_color, end_color=title_color, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    totals = [0] * 8
    
    # Data rows
    for idx, row in enumerate(data, 5):
        for col_idx, val in enumerate([row[0], row[1], row[2], row[3], row[4], row[5],
                                        row[6], row[7], row[8], row[9], row[10],
                                        row[11], row[12], row[13], row[14]], 1):
            ws.cell(row=idx, column=col_idx, value=val)
            if 7 <= col_idx <= 14:
                ws.cell(row=idx, column=col_idx).number_format = '₹#,##0'
                ws.cell(row=idx, column=col_idx).alignment = Alignment(horizontal='right')
        
        for i in range(8):
            totals[i] += float(row[6+i])
    
    # Total row
    total_row = len(data) + 5
    ws.cell(row=total_row, column=6, value="TOTAL").font = Font(bold=True)
    
    footer_color = "d1fae5" if is_weekly else "e8f4f8"
    
    for col_idx, total_val in enumerate(totals, 7):
        cell = ws.cell(row=total_row, column=col_idx, value=total_val)
        cell.font = Font(bold=True, color=title_color)
        cell.fill = PatternFill(start_color=footer_color, end_color=footer_color, fill_type="solid")
        cell.number_format = '₹#,##0'
        cell.alignment = Alignment(horizontal='right')
    
    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    for col in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col].width = 13
    ws.column_dimensions['O'].width = 10
    
    wb.save(filepath)


# -------------------------------------------------------------
# Reports Landing Page
# -------------------------------------------------------------
@reports_bp.route("/")
@login_required
def reports_home():
    return render_template("reports_home.html")


# -------------------------------------------------------------
# Universal Report Viewer
# -------------------------------------------------------------
@reports_bp.route("/view", methods=["GET", "POST"])
@login_required
def report_view():
    """Universal report viewer for ALL report types."""
    
    report_type = request.args.get("type", "").strip()
    member_id = request.args.get("member_id", "").strip()
    
    data = []
    columns = []
    start_date = ""
    end_date = ""
    due_date = ""
    center_id = ""
    centers = []
    branch_name = ""
    center_name = ""
    generation_time = ""
    
    # Get centers for monthly/weekly reports
    if report_type in ['monthly_collection', 'weekly_collection']:
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            branchid = session.get("branchid", 1)
            cursor.execute("SELECT id, center_name FROM Center WHERE branchid = ? ORDER BY center_name", (branchid,))
            centers = cursor.fetchall()
            conn.close()
        except Exception as e:
            logger.error(f"Error loading centers: {e}")
    
    if request.method == "POST":
        start_time = time.time()
        
        # Different date handling for different report types
        if report_type in ['monthly_collection', 'weekly_collection']:
            due_date = request.form.get("due_date", "").strip()
            center_id = request.form.get("center_id", "").strip()
            export_format = request.form.get("export_format", "view")
            
            if not due_date:
                flash("Please select a due date.", "warning")
                return redirect(url_for("reports.report_view", type=report_type))
        else:
            start_date = request.form.get("StartDate", "").strip()
            end_date = request.form.get("EndDate", "").strip()
            member_id = request.form.get("member_id", member_id).strip()
            
            if not start_date or not end_date:
                flash("Please select both Start Date and End Date.", "warning")
                return redirect(url_for("reports.report_view", type=report_type, member_id=member_id))
        
        if not report_type:
            flash("No report type selected.", "warning")
            return redirect(url_for("reports.reports_home"))
        
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            branchid = session.get("branchid", 1)
            
            # ══════════════════════════════════════════════════════════════
            # MONTHLY COLLECTION REPORT
            # ══════════════════════════════════════════════════════════════
            if report_type == "monthly_collection":
                cursor.execute("SELECT Name FROM Branches WHERE Id = ?", (branchid,))
                branch_info = cursor.fetchone()
                branch_name = branch_info[0] if branch_info else "Unknown Branch"
                
                center_name = "All Centers"
                if center_id:
                    cursor.execute("SELECT center_name FROM Center WHERE id = ?", (center_id,))
                    center_row = cursor.fetchone()
                    center_name = center_row[0] if center_row else "Unknown Center"
                
                query = """
                    SELECT 
                        ROW_NUMBER() OVER (ORDER BY c.center_name, m.full_name) AS SNo,
                        c.center_name, m.member_code, m.full_name, m.phone1, l.loanid,
                        ISNULL(lr.principaldueamount, 0), ISNULL(lr.interestdueamount, 0), 
                        ISNULL(lr.savingsdueamount, 0),
                        (ISNULL(lr.principaldueamount, 0) + ISNULL(lr.interestdueamount, 0) + ISNULL(lr.savingsdueamount, 0)),
                        ISNULL(lr.principalpaidamount, 0), ISNULL(lr.interestpaidamount, 0), 
                        ISNULL(lr.savingspaidamount, 0),
                        (ISNULL(lr.principalpaidamount, 0) + ISNULL(lr.interestpaidamount, 0) + ISNULL(lr.savingspaidamount, 0)),
                        CASE WHEN lr.paid = 1 THEN 'Paid' ELSE 'Pending' END
                    FROM LoanRec lr
                    JOIN Loans l ON lr.loanid = l.loanid
                    JOIN Members m ON lr.member_code = m.member_code
                    JOIN Center c ON m.center_id = c.id
                    JOIN LoanProduct lp ON l.productid = lp.ProductID
                    WHERE CAST(lr.duedate AS DATE) = ?
                      AND lp.PaymentFrequency = 'Monthly'
                      AND l.branchid = ?
                """
                
                params = [due_date, branchid]
                
                if center_id:
                    query += " AND c.id = ?"
                    params.append(center_id)
                
                query += " ORDER BY c.center_name, m.full_name"
                
                cursor.execute(query, params)
                data = cursor.fetchall()
                columns = ['S.No', 'Center', 'Member Code', 'Member Name', 'Mobile', 'Loan ID',
                          'Principal Due', 'Interest Due', 'Savings Due', 'Total Due',
                          'Principal Paid', 'Interest Paid', 'Savings Paid', 'Total Paid', 'Status']
                
                # PDF Export
                if export_format == "pdf" and PDF_AVAILABLE:
                    filename = f"Monthly_Collection_{due_date.replace('-', '')}.pdf"
                    filepath = os.path.join("reports", filename)
                    os.makedirs("reports", exist_ok=True)
                    generate_monthly_collection_pdf(due_date, branch_name, center_name, data, filepath)
                    conn.close()
                    return send_file(filepath, as_attachment=True, download_name=filename)
                
                # Excel Export
                elif export_format == "excel":
                    filename = f"Monthly_Collection_{due_date.replace('-', '')}.xlsx"
                    filepath = os.path.join("reports", filename)
                    os.makedirs("reports", exist_ok=True)
                    generate_excel(data, due_date, branch_name, center_name, filepath, is_weekly=False)
                    conn.close()
                    return send_file(filepath, as_attachment=True, download_name=filename)
            
            # ══════════════════════════════════════════════════════════════
            # WEEKLY COLLECTION REPORT
            # ══════════════════════════════════════════════════════════════
            elif report_type == "weekly_collection":
                cursor.execute("SELECT Name FROM Branches WHERE Id = ?", (branchid,))
                branch_info = cursor.fetchone()
                branch_name = branch_info[0] if branch_info else "Unknown Branch"
                
                center_name = "All Centers"
                if center_id:
                    cursor.execute("SELECT center_name FROM Center WHERE id = ?", (center_id,))
                    center_row = cursor.fetchone()
                    center_name = center_row[0] if center_row else "Unknown Center"
                
                query = """
                    SELECT 
                        ROW_NUMBER() OVER (ORDER BY c.center_name, m.full_name) AS SNo,
                        c.center_name, m.member_code, m.full_name, m.phone1, l.loanid,
                        ISNULL(lr.principaldueamount, 0), ISNULL(lr.interestdueamount, 0), 
                        ISNULL(lr.savingsdueamount, 0),
                        (ISNULL(lr.principaldueamount, 0) + ISNULL(lr.interestdueamount, 0) + ISNULL(lr.savingsdueamount, 0)),
                        ISNULL(lr.principalpaidamount, 0), ISNULL(lr.interestpaidamount, 0), 
                        ISNULL(lr.savingspaidamount, 0),
                        (ISNULL(lr.principalpaidamount, 0) + ISNULL(lr.interestpaidamount, 0) + ISNULL(lr.savingspaidamount, 0)),
                        CASE WHEN lr.paid = 1 THEN 'Paid' ELSE 'Pending' END
                    FROM LoanRec lr
                    JOIN Loans l ON lr.loanid = l.loanid
                    JOIN Members m ON lr.member_code = m.member_code
                    JOIN Center c ON m.center_id = c.id
                    JOIN LoanProduct lp ON l.productid = lp.ProductID
                    WHERE CAST(lr.duedate AS DATE) = ?
                      AND lp.PaymentFrequency = 'Weekly'
                      AND l.branchid = ?
                """
                
                params = [due_date, branchid]
                
                if center_id:
                    query += " AND c.id = ?"
                    params.append(center_id)
                
                query += " ORDER BY c.center_name, m.full_name"
                
                cursor.execute(query, params)
                data = cursor.fetchall()
                columns = ['S.No', 'Center', 'Member Code', 'Member Name', 'Mobile', 'Loan ID',
                          'Principal Due', 'Interest Due', 'Savings Due', 'Total Due',
                          'Principal Paid', 'Interest Paid', 'Savings Paid', 'Total Paid', 'Status']
                
                # PDF Export
                if export_format == "pdf" and PDF_AVAILABLE:
                    filename = f"Weekly_Collection_{due_date.replace('-', '')}.pdf"
                    filepath = os.path.join("reports", filename)
                    os.makedirs("reports", exist_ok=True)
                    generate_weekly_collection_pdf(due_date, branch_name, center_name, data, filepath)
                    conn.close()
                    return send_file(filepath, as_attachment=True, download_name=filename)
                
                # Excel Export
                elif export_format == "excel":
                    filename = f"Weekly_Collection_{due_date.replace('-', '')}.xlsx"
                    filepath = os.path.join("reports", filename)
                    os.makedirs("reports", exist_ok=True)
                    generate_excel(data, due_date, branch_name, center_name, filepath, is_weekly=True)
                    conn.close()
                    return send_file(filepath, as_attachment=True, download_name=filename)
            
            # ══════════════════════════════════════════════════════════════
            # EXISTING REPORTS (FD, RD, etc.) - NO CHANGES
            # ══════════════════════════════════════════════════════════════
            elif report_type == "fdcollections":
                query = """
                    SELECT d.FDNumber, d.MemberName, d.DepositAmount,
                           c.InterestDue, c.InterestPaid,
                           c.DepositPaid, c.PaidDate
                    FROM FDCollections c
                    JOIN FDDetails d ON c.FDDetailsId = d.Id
                    WHERE CAST(c.PaidDate AS DATE) BETWEEN ? AND ?
                    ORDER BY c.PaidDate
                """
                columns, data = fetch_data(query, (start_date, end_date))
            
            elif report_type == "rd":
                query = """
                    SELECT c.RecoveryDate, m.full_name AS Member,
                           c.Amount, ISNULL(c.Penalty, 0) AS Penalty,
                           (c.Amount + ISNULL(c.Penalty, 0)) AS TotalPaid
                    FROM RDCollections c
                    JOIN Members m ON c.MemberID = m.id
                    WHERE CAST(c.RecoveryDate AS DATE) BETWEEN ? AND ?
                    ORDER BY c.RecoveryDate
                """
                columns, data = fetch_data(query, (start_date, end_date))
            
            elif report_type == "daily":
                query = """
                    SELECT lr.createddate AS CollectionDate,
                           m.full_name AS Member,
                           lr.principalpaidamount AS Principal,
                           lr.interestpaidamount AS Interest,
                           lr.savingspaidamount AS Savings,
                           (lr.principalpaidamount + lr.interestpaidamount + lr.savingspaidamount) AS TotalPaid
                    FROM LoanRec lr
                    JOIN Members m ON lr.member_code = m.member_code
                    WHERE CAST(lr.createddate AS DATE) BETWEEN ? AND ?
                      AND (lr.principalpaidamount > 0 OR lr.interestpaidamount > 0)
                    ORDER BY lr.createddate
                """
                columns, data = fetch_data(query, (start_date, end_date))
            
            elif report_type == "fdaccounts":
                query = """
                    SELECT m.full_name AS Member,
                           f.FDNumber, f.DepositAmount,
                           f.InterestRate, f.PaymentDate,
                           f.MaturityDate, f.MaturityValue
                    FROM FDDetails f
                    JOIN Members m ON f.MemberCode = m.member_code
                    WHERE CAST(f.PaymentDate AS DATE) BETWEEN ? AND ?
                    ORDER BY f.PaymentDate
                """
                columns, data = fetch_data(query, (start_date, end_date))
            
            elif report_type == "rdaccounts":
                query = """
                    SELECT m.full_name AS Member,
                           r.RDNumber, r.MonthlyAmount,
                           r.StartDate, r.MaturityDate,
                           r.TotalAmount, r.Status
                    FROM RecurringDeposit r
                    JOIN Members m ON r.MemberCode = m.member_code
                    WHERE CAST(r.StartDate AS DATE) BETWEEN ? AND ?
                    ORDER BY r.StartDate
                """
                columns, data = fetch_data(query, (start_date, end_date))
            
            elif report_type == "fdinterest":
                query = """
                    SELECT c.PaidDate, m.full_name AS Member,
                           d.FDNumber, c.InterestDue,
                           c.InterestPaid
                    FROM FDCollections c
                    JOIN FDDetails d ON c.FDDetailsId = d.Id
                    JOIN Members m ON d.MemberCode = m.member_code
                    WHERE CAST(c.PaidDate AS DATE) BETWEEN ? AND ?
                      AND c.InterestPaid > 0
                    ORDER BY c.PaidDate
                """
                columns, data = fetch_data(query, (start_date, end_date))
            
            elif report_type == "glance":
                query = """
                    SELECT
                        CAST(d.DayendDate AS DATE)   AS ReportDate,
                        (SELECT COUNT(*) FROM Members
                         WHERE status = 'ACTIVE')    AS TotalMembers,
                        (SELECT COUNT(*) FROM Loans
                         WHERE loanstatus = 'Active') AS ActiveLoans,
                        (SELECT COUNT(*) FROM RecurringDeposit
                         WHERE Status = 'Active')    AS ActiveRDs,
                        (SELECT COUNT(*) FROM FDDetails
                         WHERE WithdrawDate IS NULL)  AS ActiveFDs
                    FROM Dayend d
                    WHERE CAST(d.DayendDate AS DATE) BETWEEN ? AND ?
                    ORDER BY d.DayendDate
                """
                columns, data = fetch_data(query, (start_date, end_date))
            
            elif report_type == "memberwise":
                if not member_id:
                    flash("Member Code is required for member-wise report.", "warning")
                    return redirect(url_for("reports.report_view", type=report_type))
                
                query = """
                    SELECT lr.duedate AS DueDate,
                           lr.principaldueamount AS PrincipalDue,
                           lr.principalpaidamount AS PrincipalPaid,
                           lr.interestdueamount AS InterestDue,
                           lr.interestpaidamount AS InterestPaid,
                           lr.savingspaidamount AS SavingsPaid,
                           CASE WHEN lr.paid = 1 THEN 'Paid' ELSE 'Pending' END AS Status
                    FROM LoanRec lr
                    WHERE lr.member_code = ?
                      AND CAST(lr.duedate AS DATE) BETWEEN ? AND ?
                    ORDER BY lr.duedate
                """
                columns, data = fetch_data(query, (member_id, start_date, end_date))
            
            else:
                flash(f"Unknown report type: '{report_type}'", "warning")
                return redirect(url_for("reports.reports_home"))
            
            conn.close()
            
            # PDF Export for date-range reports
            if request.form.get("export_pdf") and PDF_AVAILABLE and report_type not in ['monthly_collection', 'weekly_collection']:
                try:
                    filename = f"{report_type}_{start_date}_to_{end_date}.pdf"
                    filepath = os.path.join("reports", filename)
                    os.makedirs("reports", exist_ok=True)
                    
                    if report_type == "fdcollections":
                        generate_fdcollections(start_date, end_date, filepath)
                    elif report_type == "rd":
                        generate_rd_report(start_date, end_date, filepath)
                    elif report_type == "daily":
                        generate_daily_report(start_date, end_date, filepath)
                    elif report_type == "fdaccounts":
                        generate_fd_report(start_date, end_date, filepath)
                    elif report_type == "rdaccounts":
                        generate_rd_monthly_report(start_date, end_date, filepath)
                    elif report_type == "fdinterest":
                        generate_fd_interest_report(start_date, end_date, filepath)
                    elif report_type == "glance":
                        generate_glance_report(start_date, end_date, filepath)
                    elif report_type == "memberwise":
                        generate_memberwise_report(start_date, end_date, member_id, filepath)
                    
                    return send_file(filepath, as_attachment=True)
                
                except Exception as pdf_err:
                    logger.error(f"PDF generation error: {pdf_err}")
                    flash("PDF generation failed. Showing data on screen instead.", "warning")
            
            generation_time = f"{time.time() - start_time:.2f}"
        
        except Exception as e:
            logger.error(f"Report error [{report_type}]: {e}", exc_info=True)
            flash(f"Error generating report: {str(e)}", "danger")
    
    return render_template(
        "report_view.html",
        report_type=report_type,
        data=data,
        columns=columns,
        start_date=start_date,
        end_date=end_date,
        due_date=due_date,
        center_id=center_id,
        centers=centers,
        member_id=member_id,
        branch_name=branch_name,
        center_name=center_name,
        generation_time=generation_time,
        pdf_available=PDF_AVAILABLE,
    )
