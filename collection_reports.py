"""
Collection Reports Module - UDLMS
Handles monthly and weekly collection sheets with PDF/Excel export
Follows same pattern as reports.py
"""
from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for, session
from datetime import datetime
import logging
import os
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from db import get_db_connection
from login import login_required

logger = logging.getLogger(__name__)

# Try importing PDF generators
try:
    from your_pdf_generators import (
        generate_monthly_collection_pdf,
        generate_weekly_collection_pdf,
    )
    PDF_AVAILABLE = True
except ImportError as e:
    logger.warning(f"PDF generators not available: {e}")
    PDF_AVAILABLE = False

collection_reports_bp = Blueprint("collection_reports", __name__, url_prefix="/collection_reports")


# -------------------------------------------------------------
# Excel Generators
# -------------------------------------------------------------
def generate_monthly_excel(data, due_date, branch_name, center_name, filename):
    """Generate Excel for monthly collection - BLUE color scheme."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Collection"

    ws.merge_cells('A1:O1')
    ws['A1'] = f"{branch_name} - MONTHLY COLLECTION SHEET"
    ws['A1'].font = Font(size=16, bold=True, color="0c2a78")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:O2')
    ws['A2'] = f"Due Date: {datetime.strptime(due_date, '%Y-%m-%d').strftime('%d-%b-%Y')} | Center: {center_name}"
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['S.No', 'Center', 'Member Code', 'Member Name', 'Mobile', 'Loan ID',
               'Principal Due', 'Interest Due', 'Savings Due', 'Total Due',
               'Principal Paid', 'Interest Paid', 'Savings Paid', 'Total Paid', 'Status']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="0c2a78", end_color="0c2a78", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    totals = [0] * 8

    for idx, row in enumerate(data, 5):
        for col_idx, val in enumerate([row[0], row[1], row[2], row[3], row[4], row[5],
                                        row[6], row[7], row[8], row[9], row[10],
                                        row[11], row[12], row[13], row[14]], 1):
            ws.cell(row=idx, column=col_idx, value=val)
            if 7 <= col_idx <= 14:
                ws.cell(row=idx, column=col_idx).number_format = '₹#,##0'
                ws.cell(row=idx, column=col_idx).alignment = Alignment(horizontal='right')

        for i in range(8):
            totals[i] += float(row[6+i])

    total_row = len(data) + 5
    ws.cell(row=total_row, column=6, value="TOTAL").font = Font(bold=True)

    for col_idx, total_val in enumerate(totals, 7):
        cell = ws.cell(row=total_row, column=col_idx, value=total_val)
        cell.font = Font(bold=True, color="0c2a78")
        cell.fill = PatternFill(start_color="e8f4f8", end_color="e8f4f8", fill_type="solid")
        cell.number_format = '₹#,##0'
        cell.alignment = Alignment(horizontal='right')

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    for col in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col].width = 13
    ws.column_dimensions['O'].width = 10

    wb.save(filename)


def generate_weekly_excel(data, due_date, branch_name, center_name, filename):
    """Generate Excel for weekly collection - GREEN color scheme."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Collection"

    ws.merge_cells('A1:O1')
    ws['A1'] = f"{branch_name} - WEEKLY COLLECTION SHEET"
    ws['A1'].font = Font(size=16, bold=True, color="059669")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:O2')
    ws['A2'] = f"Due Date: {datetime.strptime(due_date, '%Y-%m-%d').strftime('%d-%b-%Y')} | Center: {center_name}"
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['S.No', 'Center', 'Member Code', 'Member Name', 'Mobile', 'Loan ID',
               'Principal Due', 'Interest Due', 'Savings Due', 'Total Due',
               'Principal Paid', 'Interest Paid', 'Savings Paid', 'Total Paid', 'Status']

    # GREEN header
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    totals = [0] * 8

    for idx, row in enumerate(data, 5):
        for col_idx, val in enumerate([row[0], row[1], row[2], row[3], row[4], row[5],
                                        row[6], row[7], row[8], row[9], row[10],
                                        row[11], row[12], row[13], row[14]], 1):
            ws.cell(row=idx, column=col_idx, value=val)
            if 7 <= col_idx <= 14:
                ws.cell(row=idx, column=col_idx).number_format = '₹#,##0'
                ws.cell(row=idx, column=col_idx).alignment = Alignment(horizontal='right')

        for i in range(8):
            totals[i] += float(row[6+i])

    total_row = len(data) + 5
    ws.cell(row=total_row, column=6, value="TOTAL").font = Font(bold=True)

    # GREEN total row
    for col_idx, total_val in enumerate(totals, 7):
        cell = ws.cell(row=total_row, column=col_idx, value=total_val)
        cell.font = Font(bold=True, color="059669")
        cell.fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
        cell.number_format = '₹#,##0'
        cell.alignment = Alignment(horizontal='right')

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    for col in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col].width = 13
    ws.column_dimensions['O'].width = 10

    wb.save(filename)


# -------------------------------------------------------------
# MONTHLY COLLECTION SHEET
# -------------------------------------------------------------
@collection_reports_bp.route("/monthly", methods=["GET", "POST"])
@login_required
def monthly_collection():
    """Monthly collection sheet."""

    start_time = time.time()

    # Get centers for dropdown
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        branchid = session.get("branchid", 1)
        cursor.execute("SELECT id, center_name FROM Center WHERE branchid = ? ORDER BY center_name", (branchid,))
        centers = cursor.fetchall()
        conn.close()
    except Exception as e:
        logger.error(f"Error loading centers: {e}")
        centers = []

    if request.method == "POST":
        due_date = request.form.get("due_date", "").strip()
        center_id = request.form.get("center_id", "").strip()
        export_format = request.form.get("export_format", "view")

        if not due_date:
            flash("Please select a due date.", "warning")
            return render_template("collection_reports/monthly_collection.html", centers=centers)

        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            # Get branch info
            branchid = session.get("branchid", 1)
            cursor.execute("SELECT Name FROM Branches WHERE Id = ?", (branchid,))
            branch_info = cursor.fetchone()
            branch_name = branch_info[0] if branch_info else "Unknown Branch"

            # Get center name
            center_name = "All Centers"
            if center_id:
                cursor.execute("SELECT center_name FROM Center WHERE id = ?", (center_id,))
                center_row = cursor.fetchone()
                center_name = center_row[0] if center_row else "Unknown Center"

            # Query with correct schema (phone1, center_id, loanid, PaymentFrequency)
            query = """
                SELECT 
                    ROW_NUMBER() OVER (ORDER BY c.center_name, m.full_name) AS SNo,
                    c.center_name AS Center,
                    m.member_code AS MemberCode,
                    m.full_name AS MemberName,
                    m.phone1 AS Mobile,
                    l.loanid AS LoanID,
                    ISNULL(lr.principaldueamount, 0) AS PrincipalDue,
                    ISNULL(lr.interestdueamount, 0) AS InterestDue,
                    ISNULL(lr.savingsdueamount, 0) AS SavingsDue,
                    (ISNULL(lr.principaldueamount, 0) + ISNULL(lr.interestdueamount, 0) + ISNULL(lr.savingsdueamount, 0)) AS TotalDue,
                    ISNULL(lr.principalpaidamount, 0) AS PrincipalPaid,
                    ISNULL(lr.interestpaidamount, 0) AS InterestPaid,
                    ISNULL(lr.savingspaidamount, 0) AS SavingsPaid,
                    (ISNULL(lr.principalpaidamount, 0) + ISNULL(lr.interestpaidamount, 0) + ISNULL(lr.savingspaidamount, 0)) AS TotalPaid,
                    CASE WHEN lr.paid = 1 THEN 'Paid' ELSE 'Pending' END AS Status
                FROM LoanRec lr
                JOIN Loans l ON lr.loanid = l.loanid
                JOIN Members m ON lr.member_code = m.member_code
                JOIN Center c ON m.center_id = c.id
                JOIN LoanProduct lp ON l.productid = lp.ProductID
                WHERE CAST(lr.duedate AS DATE) = ?
                  AND lp.PaymentFrequency = 'Monthly'
                  AND l.branchid = ?
            """

            params = [due_date, branchid]

            if center_id:
                query += " AND c.id = ?"
                params.append(center_id)

            query += " ORDER BY c.center_name, m.full_name"

            cursor.execute(query, params)
            data = cursor.fetchall()
            conn.close()

            total_time = time.time() - start_time
            logger.info(f"Monthly collection: {len(data)} records in {total_time:.2f}s")

            # PDF Export
            if export_format == "pdf":
                try:
                    if not PDF_AVAILABLE:
                        flash("PDF export not available. Please install reportlab.", "warning")
                        return redirect(url_for("collection_reports.monthly_collection"))

                    filename = f"Monthly_Collection_{due_date.replace('-', '')}.pdf"
                    filepath = os.path.join("reports", filename)
                    os.makedirs("reports", exist_ok=True)

                    generate_monthly_collection_pdf(due_date, branch_name, center_name, data, filepath)
                    logger.info(f"PDF created: {filepath}")

                    return send_file(filepath, as_attachment=True, download_name=filename)

                except Exception as pdf_err:
                    logger.error(f"PDF generation error: {pdf_err}")
                    flash("PDF generation failed. Showing data on screen instead.", "warning")

            # Excel Export
            elif export_format == "excel":
                try:
                    filename = f"Monthly_Collection_{due_date.replace('-', '')}.xlsx"
                    filepath = os.path.join("reports", filename)
                    os.makedirs("reports", exist_ok=True)

                    generate_monthly_excel(data, due_date, branch_name, center_name, filepath)
                    logger.info(f"Excel created: {filepath}")

                    return send_file(filepath, as_attachment=True, download_name=filename)

                except Exception as excel_err:
                    logger.error(f"Excel generation error: {excel_err}")
                    flash("Excel generation failed. Showing data on screen instead.", "warning")

            # View on screen
            return render_template("collection_reports/monthly_collection.html",
                                 data=data,
                                 due_date=due_date,
                                 centers=centers,
                                 selected_center=center_id,
                                 branch_name=branch_name,
                                 center_name=center_name,
                                 generation_time=f"{total_time:.2f}",
                                 pdf_available=PDF_AVAILABLE)

        except Exception as e:
            logger.error(f"Monthly collection error: {e}", exc_info=True)
            flash(f"Error generating report: {str(e)}", "danger")
            return render_template("collection_reports/monthly_collection.html", centers=centers)

    return render_template("collection_reports/monthly_collection.html", centers=centers, pdf_available=PDF_AVAILABLE)


# -------------------------------------------------------------
# WEEKLY COLLECTION SHEET
# -------------------------------------------------------------
@collection_reports_bp.route("/weekly", methods=["GET", "POST"])
@login_required
def weekly_collection():
    """Weekly collection sheet."""

    start_time = time.time()

    # Get centers for dropdown
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        branchid = session.get("branchid", 1)
        cursor.execute("SELECT id, center_name FROM Center WHERE branchid = ? ORDER BY center_name", (branchid,))
        centers = cursor.fetchall()
        conn.close()
    except Exception as e:
        logger.error(f"Error loading centers: {e}")
        centers = []

    if request.method == "POST":
        due_date = request.form.get("due_date", "").strip()
        center_id = request.form.get("center_id", "").strip()
        export_format = request.form.get("export_format", "view")

        if not due_date:
            flash("Please select a due date.", "warning")
            return render_template("collection_reports/weekly_collection.html", centers=centers)

        try:
            conn = get_db_connection()
            cursor = conn.cursor()

            # Get branch info
            branchid = session.get("branchid", 1)
            cursor.execute("SELECT Name FROM Branches WHERE Id = ?", (branchid,))
            branch_info = cursor.fetchone()
            branch_name = branch_info[0] if branch_info else "Unknown Branch"

            # Get center name
            center_name = "All Centers"
            if center_id:
                cursor.execute("SELECT center_name FROM Center WHERE id = ?", (center_id,))
                center_row = cursor.fetchone()
                center_name = center_row[0] if center_row else "Unknown Center"

            # Query - WEEKLY PaymentFrequency
            query = """
                SELECT 
                    ROW_NUMBER() OVER (ORDER BY c.center_name, m.full_name) AS SNo,
                    c.center_name, m.member_code, m.full_name, m.phone1, l.loanid,
                    ISNULL(lr.principaldueamount, 0), ISNULL(lr.interestdueamount, 0), ISNULL(lr.savingsdueamount, 0),
                    (ISNULL(lr.principaldueamount, 0) + ISNULL(lr.interestdueamount, 0) + ISNULL(lr.savingsdueamount, 0)),
                    ISNULL(lr.principalpaidamount, 0), ISNULL(lr.interestpaidamount, 0), ISNULL(lr.savingspaidamount, 0),
                    (ISNULL(lr.principalpaidamount, 0) + ISNULL(lr.interestpaidamount, 0) + ISNULL(lr.savingspaidamount, 0)),
                    CASE WHEN lr.paid = 1 THEN 'Paid' ELSE 'Pending' END
                FROM LoanRec lr
                JOIN Loans l ON lr.loanid = l.loanid
                JOIN Members m ON lr.member_code = m.member_code
                JOIN Center c ON m.center_id = c.id
                JOIN LoanProduct lp ON l.productid = lp.ProductID
                WHERE CAST(lr.duedate AS DATE) = ?
                  AND lp.PaymentFrequency = 'Weekly'
                  AND l.branchid = ?
            """

            params = [due_date, branchid]

            if center_id:
                query += " AND c.id = ?"
                params.append(center_id)

            query += " ORDER BY c.center_name, m.full_name"

            cursor.execute(query, params)
            data = cursor.fetchall()
            conn.close()

            total_time = time.time() - start_time
            logger.info(f"Weekly collection: {len(data)} records in {total_time:.2f}s")

            # PDF Export
            if export_format == "pdf":
                try:
                    if not PDF_AVAILABLE:
                        flash("PDF export not available. Please install reportlab.", "warning")
                        return redirect(url_for("collection_reports.weekly_collection"))

                    filename = f"Weekly_Collection_{due_date.replace('-', '')}.pdf"
                    filepath = os.path.join("reports", filename)
                    os.makedirs("reports", exist_ok=True)

                    generate_weekly_collection_pdf(due_date, branch_name, center_name, data, filepath)
                    logger.info(f"PDF created: {filepath}")

                    return send_file(filepath, as_attachment=True, download_name=filename)

                except Exception as pdf_err:
                    logger.error(f"PDF generation error: {pdf_err}")
                    flash("PDF generation failed. Showing data on screen instead.", "warning")

            # Excel Export
            elif export_format == "excel":
                try:
                    filename = f"Weekly_Collection_{due_date.replace('-', '')}.xlsx"
                    filepath = os.path.join("reports", filename)
                    os.makedirs("reports", exist_ok=True)

                    generate_weekly_excel(data, due_date, branch_name, center_name, filepath)
                    logger.info(f"Excel created: {filepath}")

                    return send_file(filepath, as_attachment=True, download_name=filename)

                except Exception as excel_err:
                    logger.error(f"Excel generation error: {excel_err}")
                    flash("Excel generation failed. Showing data on screen instead.", "warning")

            # View on screen
            return render_template("collection_reports/weekly_collection.html",
                                 data=data,
                                 due_date=due_date,
                                 centers=centers,
                                 selected_center=center_id,
                                 branch_name=branch_name,
                                 center_name=center_name,
                                 generation_time=f"{total_time:.2f}",
                                 pdf_available=PDF_AVAILABLE)

        except Exception as e:
            logger.error(f"Weekly collection error: {e}", exc_info=True)
            flash(f"Error generating report: {str(e)}", "danger")
            return render_template("collection_reports/weekly_collection.html", centers=centers)

    return render_template("collection_reports/weekly_collection.html", centers=centers, pdf_available=PDF_AVAILABLE)